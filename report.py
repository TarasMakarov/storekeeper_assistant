from datetime import datetime
from openpyxl.styles import Side, Border


# prepare file for preview
def prepare_file(file, maximum):
    file.active.delete_cols(1)
    file.active.delete_rows(5)
    # удаление позиций, отсутствующих на складе хранения
    # отбор позиций для пополнения склада отгрузки
    row = 6
    col = 1
    sku = file.active.cell(row=row, column=col).value
    while isinstance(sku, str):
        amount_store = file.active.cell(row=row, column=3).value  # кол-во на складе хранения
        amount_sell = file.active.cell(row=row, column=4).value  # кол-во на складе отгрузки
        if amount_sell is None:
            amount_sell = 0
        if amount_store is None or amount_sell > maximum:
            file.active.delete_rows(row)
            row -= 1
        row += 1
        sku = file.active.cell(row=row, column=col).value
    # ширина столбцов, отрисовка границ ячеек
    for i in 'AEF':
        file.active.column_dimensions[i].width = 12
    file.active.column_dimensions['B'].width = 50
    file.active.column_dimensions['C'].width = 9
    file.active.column_dimensions['D'].width = 9
    rows = file.active.max_row - 5
    row = 5
    side = Side(border_style="thin", color="000000")
    for _ in range(rows):
        file.active.row_dimensions[row].height = 15
        file.active.cell(row=row, column=5).border = Border(left=side, right=side, top=side,
                                                            bottom=side)
        file.active.cell(row=row, column=6).border = Border(left=side, right=side, top=side,
                                                            bottom=side)
        row += 1
    return file


def create_cells_list(cells_list, cells_set):
    cells_list = cells_list.split(', ')
    for cell in cells_list:
        if cell not in cells_set:
            cells_set.append(cell)
    return cells_set


def fill_report(file, file_balance_batch):
    # файл 'остатки АТ по партиям'
    cells_file = file_balance_batch

    # список sku для пополнения склада отгрузки
    sku_list = []
    row_405 = 6
    sku_405 = file.active.cell(row_405, 1).value
    while isinstance(sku_405, str):
        sku_list.append(sku_405)
        row_405 += 1
        sku_405 = file.active.cell(row_405, 1).value

    sku_amount = len(sku_list)
    row_406 = 4
    # sku_report - код товара, который надо пополнить
    # sku_batch_curr - код товара (рассматриваемый) в файле "Остатки по партиям"
    # sku_batch_next - код товара (следующий) в файле "Остатки по партиям"
    # sku_store_405 - ячейка на складе хранения (откуда взять)
    # sku_store_406 - ячейка на складе отгрузки (куда поставить)
    # cells_405 - ячейки на складе хранения
    # cells_406 - ячейки на складе отгрузки
    # row_406 - номер строки в файле "Остатки по партиям"
    cells_405_set = []
    cells_406_set = []
    sku_batch_curr = cells_file.active.cell(row_406, 4).value
    while sku_batch_curr is not None:
        sku_batch_curr = sku_batch_curr.split('.')[0]
        if sku_batch_curr not in sku_list:
            row_406 += 1
            sku_batch_curr = cells_file.active.cell(row_406, 4).value
        else:
            cells_405_list = cells_file.active.cell(row_406, 1).value
            if cells_405_list is not None:
                cells_405_set = create_cells_list(cells_405_list, cells_405_set)
            cells_406_list = cells_file.active.cell(row_406, 2).value
            if cells_406_list is not None:
                cells_406_set = create_cells_list(cells_406_list, cells_406_set)
            sku_batch_next = cells_file.active.cell(row_406 + 1, 4).value
            if sku_batch_next is not None:
                sku_batch_next = sku_batch_next.split('.')[0]
                if sku_batch_next != sku_batch_curr:
                    row_report = 6
                    for _ in range(sku_amount):
                        if sku_batch_curr == file.active.cell(row_report, 1).value:
                            if not cells_405_set:
                                cells_405_set = ['пусто']
                            file.active.cell(row_report, 5).value = cells_405_set[0]
                            if not cells_406_set:
                                cells_406_set = ['пусто']
                            file.active.cell(row_report, 6).value = ", ".join(cells_406_set)
                            cells_405_set = []
                            cells_406_set = []
                            break
                        else:
                            row_report += 1
            else:
                row_report = 6
                for _ in range(sku_amount):
                    if sku_batch_curr == file.active.cell(row_report, 1).value:
                        file.active.cell(row_report, 5).value = cells_405_set[0]
                        file.active.cell(row_report, 6).value = ", ".join(cells_406_set)
                        cells_405_set = []
                        cells_406_set = []
                        break
                    else:
                        row_report += 1
            sku_batch_curr = sku_batch_next
            row_406 += 1
    return file


def create_report(report, file_balance_batch, maximum, warehouse):
    if warehouse == '406':
        report.active.delete_cols(19, 16)
        report.active.delete_cols(4, 13)
    else:
        report.active.delete_cols(21, 14)
        report.active.delete_cols(4, 15)

    report = prepare_file(report, maximum)
    report = fill_report(report, file_balance_batch)
    report.save(f'Отчет от {datetime.today().strftime("%d.%m.%y %H_%M_%S")} ({warehouse}).xlsx')
