import tkinter as tk
from datetime import datetime
from tkinter import filedialog as fd
from tkinter.ttk import Notebook, Combobox
from tkcalendar import DateEntry
from tkinter import messagebox as mb
from babel import numbers  # DON'T DELETE - for auto-py-to-exe

import openpyxl as op

import mistake
import report
import warehouses
import inventory


class Application(tk.Tk):

    def __init__(self):
        super().__init__()
        self.title('АРОМА СПб')
        screenwidth = self.winfo_screenwidth()
        screenheight = self.winfo_screenheight()
        app_win_width = 700
        app_win_height = 500
        x = (screenwidth - app_win_width) / 2
        y = (screenheight - app_win_height) / 2
        self.geometry("%dx%d+%d+%d" % (app_win_width, app_win_height, x, y))

        self.tabs_control = Notebook(self)
        self.tabs_control.enable_traversal()

        self.storekeeper = tk.Frame(self.tabs_control)
        self.tabs_control.add(self.storekeeper, text='Помощник кладовщика')
        self.tabs_control.pack(fill='both', expand=True)

        self.foreman = tk.Frame(self.tabs_control)
        self.tabs_control.add(self.foreman, text='Помощник старшего смены')
        self.tabs_control.pack(fill='both', expand=True)

        self.file_name_bb = ''
        self.file_name_cb = ''

        self.put_widgets_storekeeper()
        self.put_widgets_foreman()

    def put_widgets_storekeeper(self):  # storekeeper

        OPERATION_WAS_NOT_PERFORMED = {'Данная операция не выполнялась.': ''}
        self.MISTAKES_NOT_FOUND = {'Ошибки не обнаружены': ''}

        self.entry_file_path = tk.Entry(self.storekeeper)
        self.entry_file_path.insert(0, 'Вставьте файл отчета "Остатки АТ по партиям"')
        self.entry_file_path['width'] = 75
        self.entry_file_path.place(x=10, y=14)

        self.button_choose_file = tk.Button(self.storekeeper)
        self.button_choose_file.place(x=500, y=10)
        self.button_choose_file['text'] = 'Выберите файл'
        self.button_choose_file['font'] = 'Arial', 11
        self.button_choose_file['command'] = self.choose_file

        self.header = tk.Label(self.storekeeper)
        self.header['text'] = 'Провести следующие операции:'
        self.header['font'] = 'Arial', 12, 'italic'
        self.header.place(x=50, y=50)

        self.check_button_search_sku_in_two_places_value = tk.IntVar()
        self.check_button_search_sku_in_two_places = tk.Checkbutton(
            self.storekeeper,
            variable=self.check_button_search_sku_in_two_places_value,
            offvalue=0,
            onvalue=1)
        self.check_button_search_sku_in_two_places[
            'text'] = '  Найти товар, размещенный в двух и более ячейках отгрузки'
        self.check_button_search_sku_in_two_places.place(x=10, y=80)
        self.duplicates_result = OPERATION_WAS_NOT_PERFORMED

        self.check_button_search_two_or_more_sku_in_one_place_value = tk.IntVar()
        self.check_button_search_two_or_more_sku_in_one_place = tk.Checkbutton(
            self.storekeeper,
            variable=self.check_button_search_two_or_more_sku_in_one_place_value,
            offvalue=0,
            onvalue=1)
        self.check_button_search_two_or_more_sku_in_one_place[
            'text'] = '  Найти ячейки, в которых больше одного вида товара'
        self.check_button_search_two_or_more_sku_in_one_place.place(x=10, y=110)
        self.wrong_place = OPERATION_WAS_NOT_PERFORMED

        self.check_search_cells_empty_value = tk.IntVar()
        self.check_search_cells_empty = tk.Checkbutton(
            self.storekeeper,
            variable=self.check_search_cells_empty_value,
            offvalue=0,
            onvalue=1)
        self.check_search_cells_empty['text'] = '  Найти пустые нижние ячейки на складах отгрузки'
        self.check_search_cells_empty.place(x=10, y=140)
        self.cells_empty = OPERATION_WAS_NOT_PERFORMED

        self.header_2 = tk.Label(self.storekeeper)
        self.header_2['text'] = 'Если хотите сформировать отчет для пополнения склада отгрузки, \n поставьте галочку ' \
                                'и выберите файл отчета "Текущие остатки":'
        self.header_2['font'] = 'Arial', 12, 'italic'
        self.header_2.place(x=50, y=170)

        self.check_button_fill_storage_value = tk.IntVar()
        self.check_button_fill_storage = tk.Checkbutton(
            self.storekeeper,
            variable=self.check_button_fill_storage_value,
            offvalue=0,
            onvalue=1)
        self.check_button_fill_storage[
            'text'] = 'Сформировать отчет для пополнения склада отгрузки'
        self.check_button_fill_storage.place(x=10, y=220)

        self.unit_max = tk.StringVar()
        self.maximum = tk.Entry(self.storekeeper, textvariable=self.unit_max)
        self.maximum['width'] = 15
        self.maximum.place(x=10, y=250)

        self.header_max = tk.Label(self.storekeeper)
        self.header_max['text'] = ' Впишите максимальный остаток товара на складе отгрузки'
        self.header_max.place(x=50, y=250)

        self.entry_file_current_balance = tk.Entry(self.storekeeper)
        self.entry_file_current_balance.insert(0, 'Вставьте файл отчета "Текущие остатки (цена, шт., бр., пост.)"')
        self.entry_file_current_balance['width'] = 75
        self.entry_file_current_balance.place(x=10, y=300)

        self.button_choose_file_cb = tk.Button(self.storekeeper)
        self.button_choose_file_cb.place(x=500, y=296)
        self.button_choose_file_cb['text'] = 'Выберите файл'
        self.button_choose_file_cb['font'] = 'Arial', 11
        self.button_choose_file_cb['command'] = self.choose_file_cb

        self.button_run_actions = tk.Button(self.storekeeper)
        self.button_run_actions['text'] = 'Начать'
        self.button_run_actions['font'] = 'Arial', 11
        self.button_run_actions['command'] = self.run_actions
        self.button_run_actions.place(x=50, y=350)

    # Выбор файла 'Остатки АТ по партиям' storekeeper
    def choose_file(self):
        self.file_name_bb = fd.askopenfilename(filetypes=(('Excel files', '.xlsx'),))
        self.entry_file_path.delete(0, 'end')
        self.entry_file_path.insert(0, self.file_name_bb)

    # Выбор файла 'Текущие остатки' storekeeper
    def choose_file_cb(self):
        self.file_name_cb = fd.askopenfilename(filetypes=(("Excel files", ".xlsx"),))
        self.entry_file_current_balance.delete(0, 'end')
        self.entry_file_current_balance.insert(0, self.file_name_cb)

    def open_file(self, file_name):  # storekeeper
        sheet = op.load_workbook(file_name).sheetnames[0]
        if sheet == 'tmp':
            self.file_balance_current = op.load_workbook(file_name)  # Открытие файла 'Текущие остатки'
            self.file_path_cb = file_name
        else:
            self.file_balance_batch = op.load_workbook(file_name)  # Открытие файла 'Остатки АТ по партиям

    def run_actions(self):  # storekeeper
        while self.file_name_bb == '':
            mb.showwarning(title='Выбор файла', message='Выберите файл "Остатки АТ по партиям"')
            self.choose_file()
            return
        self.open_file(self.file_name_bb)  # Остатки АТ по партиям
        if self.check_button_fill_storage_value.get():
            while self.file_name_cb == '':
                mb.showwarning(title='Выбор файла', message='Выберите файл "Текущие остатки (цена, шт., бр., пост.)"')
                self.choose_file_cb()
            self.open_file(self.file_name_cb)  # Текущие остатки
        if self.check_button_fill_storage_value.get():
            try:
                maximum = int(self.maximum.get())
            except ValueError:
                mb.showwarning(title='Поле должно быть заполнено', message='Впишите максимальный остаток товара на складе отгрузки')
                return
            report.create_report(self.file_balance_current, self.file_balance_batch, maximum, warehouses.warehouse_406)
            report.create_report(self.file_balance_current, self.file_balance_batch, maximum, warehouses.warehouse_437)
        if self.check_button_search_sku_in_two_places_value.get():
            self.duplicates_result = mistake.search_sku_in_two_places(self.file_balance_batch)
        if self.check_button_search_two_or_more_sku_in_one_place_value.get():
            self.wrong_place = mistake.search_two_or_more_sku_in_one_place(self.file_balance_batch)
        if self.check_search_cells_empty_value.get():
            self.cells_empty = mistake.search_cells_empty(self.file_balance_batch)
        self.fill_mistakes()
        self.destroy()

    def fill_mistakes(self):  # storekeeper
        row = 1
        col = 1
        result = op.Workbook()
        result.remove(result.active)
        sheet_1 = result.create_sheet('1')
        sheet_1.cell(row=row, column=col).value = 'Товар, размещенный в двух и более ячейках отгрузки'
        if not self.duplicates_result:
            self.duplicates_result = self.MISTAKES_NOT_FOUND
        for k, v in self.duplicates_result.items():
            row += 1
            sheet_1.cell(row=row, column=col).value = k
            sheet_1.cell(row=row, column=col + 1).value = ', '.join(v)

        row += 2
        sheet_1.cell(row=row, column=col).value = 'Ячейки, в которых больше одного вида товара'
        if not self.wrong_place:
            self.wrong_place = self.MISTAKES_NOT_FOUND
        for key, value in self.wrong_place.items():
            row += 1
            sheet_1.cell(row=row, column=col).value = key
            sheet_1.cell(row=row, column=col + 1).value = value

        row += 2
        sheet_1.cell(row=row, column=col).value = 'Свободные ячейки'
        row += 1
        sheet_1.cell(row=row, column=col).value = 406
        sheet_1.cell(row=row, column=col + 1).value = 437
        row_406 = row_437 = row + 1
        if self.check_search_cells_empty_value.get() == 0:
            sheet_1.cell(row=row_406, column=col).value = next(iter(self.cells_empty))
        else:
            for c_406 in self.cells_empty[0]:
                sheet_1.cell(row=row_406, column=col).value = c_406
                row_406 += 1
            for c_437 in self.cells_empty[1]:
                sheet_1.cell(row=row_437, column=col + 1).value = c_437
                row_437 += 1

        result.save(f'Ошибки от {datetime.today().strftime("%d.%m.%y %H_%M_%S")}.xlsx')

    # Выпадающий список ячеек зависит от номера выбранного склада  # foreman
    def select_cells(self, *args):
        self.box_cell_start.set('')
        self.box_cell_finish.set('')
        self.cells_in_combobox = self.warehouses_and_cells[self.wh.get()]  # все ячейки определенного склада
        self.box_cell_start['values'] = self.cells_in_combobox
        self.box_cell_finish['values'] = self.cells_in_combobox
        self.box_cell_start.current(0)
        self.box_cell_finish.current(0)

    def put_widgets_foreman(self):  # foreman

        self.warehouses_and_cells = {
            warehouses.warehouse_405: warehouses.cells_all_405, warehouses.warehouse_406: warehouses.cells_all_406,
            warehouses.warehouse_436: warehouses.cells_all_436, warehouses.warehouse_437: warehouses.cells_all_437}

        self.check_button_inventory_cells_report_value = tk.IntVar()
        self.check_button_inventory_cells_report = tk.Checkbutton(
            self.foreman,
            variable=self.check_button_inventory_cells_report_value,
            offvalue=0,
            onvalue=1)
        self.check_button_inventory_cells_report[
            'text'] = '  Создать инвентарную ведомость по ячейкам'
        self.check_button_inventory_cells_report.place(x=10, y=15)

        self.header_inventory_cells = tk.Label(self.foreman)
        self.header_inventory_cells['text'] = 'Выбрать ячейки для инвентаризации:'
        self.header_inventory_cells['font'] = 'Arial', 12, 'italic'
        self.header_inventory_cells.place(x=50, y=50)

        self.label_wh = tk.Label(self.foreman, text='Склад №')
        self.label_wh.place(x=10, y=80)

        self.wh = tk.StringVar()

        self.box_warehouses = Combobox(self.foreman, values=[warehouses.warehouses[1], warehouses.warehouses[3]],
                                       textvariable=self.wh, state='readonly')
        self.box_warehouses.current(0)
        self.box_warehouses['width'] = 7
        self.box_warehouses.place(x=65, y=80)

        self.wh.trace('w', self.select_cells)

        self.cells_in_combobox = warehouses.cells_all_406

        self.label_cell_start = tk.Label(self.foreman, text='Начать с ячейки')
        self.label_cell_start.place(x=160, y=80)

        self.box_cell_start = Combobox(self.foreman, values=self.cells_in_combobox, state='readonly')
        self.box_cell_start.current(0)
        self.box_cell_start['width'] = 12
        self.box_cell_start.place(x=260, y=80)

        self.label_cell_finish = tk.Label(self.foreman, text='Закончить ячейкой')
        self.label_cell_finish.place(x=385, y=80)

        self.box_cell_finish = Combobox(self.foreman, values=self.cells_in_combobox, state='readonly')
        self.box_cell_finish.current(0)
        self.box_cell_finish['width'] = 12
        self.box_cell_finish.place(x=500, y=80)

        self.entry_file_path_AT = tk.Entry(self.foreman)
        self.entry_file_path_AT.insert(0, 'Вставьте файл отчета "Остатки АТ по партиям"')
        self.entry_file_path_AT['width'] = 75
        self.entry_file_path_AT.place(x=10, y=126)

        self.button_choose_file_AT = tk.Button(self.foreman)
        self.button_choose_file_AT.place(x=500, y=120)
        self.button_choose_file_AT['text'] = 'Выберите файл'
        self.button_choose_file_AT['font'] = 'Arial', 11
        self.button_choose_file_AT['command'] = self.choose_file_AT

        self.check_button_inventory_supervisor_report_value = tk.IntVar()
        self.check_button_inventory_supervisor_report = tk.Checkbutton(
            self.foreman,
            variable=self.check_button_inventory_supervisor_report_value,
            offvalue=0,
            onvalue=1)
        self.check_button_inventory_supervisor_report[
            'text'] = '  Создать инвентарную ведомость по ошибкам'
        self.check_button_inventory_supervisor_report.place(x=10, y=175)

        self.header_inventory_supervisor = tk.Label(self.foreman)
        self.header_inventory_supervisor['text'] = 'Выбрать даты для инвентаризации:'
        self.header_inventory_supervisor['font'] = 'Arial', 12, 'italic'
        self.header_inventory_supervisor.place(x=50, y=200)

        self.label_date_start = tk.Label(self.foreman, text='С')
        self.label_date_start.place(x=20, y=230)

        self.entry_date_start = DateEntry(self.foreman, date_pattern='dd-mm-YYYY')
        self.entry_date_start.place(x=35, y=230)

        self.label_date_finish = tk.Label(self.foreman, text='По')
        self.label_date_finish.place(x=150, y=230)

        self.entry_date_finish = DateEntry(self.foreman, date_pattern='dd-mm-YYYY')
        self.entry_date_finish.place(x=170, y=230)

        self.entry_file_path_supervisor = tk.Entry(self.foreman)
        self.entry_file_path_supervisor.insert(0, 'Вставьте файл с ошибками')
        self.entry_file_path_supervisor['width'] = 75
        self.entry_file_path_supervisor.place(x=10, y=276)

        self.button_choose_file_supervisor = tk.Button(self.foreman)
        self.button_choose_file_supervisor.place(x=500, y=270)
        self.button_choose_file_supervisor['text'] = 'Выберите файл'
        self.button_choose_file_supervisor['font'] = 'Arial', 11
        self.button_choose_file_supervisor['command'] = self.choose_file_supervisor

        self.button_run_inventory = tk.Button(self.foreman)
        self.button_run_inventory['text'] = 'Начать'
        self.button_run_inventory['font'] = 'Arial', 11
        self.button_run_inventory['command'] = self.run_inventory
        self.button_run_inventory.place(x=500, y=350)

    def choose_file_AT(self):
        self.file_name_AT = fd.askopenfilename(filetypes=(('Excel files', '.xlsx'),))
        self.entry_file_path_AT.delete(0, 'end')
        self.entry_file_path_AT.insert(0, self.file_name_AT)

    def choose_file_supervisor(self):
        self.file_name_supervisor = fd.askopenfilename(filetypes=(('Excel files', '.xlsx'),))
        self.entry_file_path_supervisor.delete(0, 'end')
        self.entry_file_path_supervisor.insert(0, self.file_name_supervisor)

    def open_file_for_foreman(self, file_name):  # foreman
        sheets = op.load_workbook(file_name).sheetnames
        if 'СБОРКА' in sheets:
            self.file_supervisor = op.load_workbook(file_name)  # Открытие файла 'РАБОЧАЯ'
        else:
            self.file_AT = op.load_workbook(file_name)  # Открытие файла 'Остатки АТ по партиям

    def run_inventory(self):
        self.open_file_for_foreman(self.file_name_AT)
        self.open_file_for_foreman(self.file_name_supervisor)
        if self.check_button_inventory_cells_report_value.get():
            inventory.create_file_inventory_cells(self.file_AT, self.box_warehouses.get(), self.box_cell_start.get(),
                                                  self.box_cell_finish.get())
        if self.check_button_inventory_supervisor_report_value.get():
            inventory.create_file_inventory_supervisor(self.file_AT, self.file_supervisor, self.box_warehouses.get(),
                                                       self.entry_date_start.get_date(),
                                                       self.entry_date_finish.get_date())


app = Application()
app.mainloop()
