"""
В этом файле содержаться:
 - функции создания шаблонов отчетов
 - функции подготовки к печати создаваемых файлов
"""

from datetime import datetime
from openpyxl.worksheet.page import PageMargins
import openpyxl as op

import design_of_report as dor


def create_report_inventory_template(warehouse, item):
    report_inventory_template = op.Workbook()
    report_inventory_template.remove(report_inventory_template.active)
    report_inventory_template.create_sheet(title=warehouse)

    report_inventory_template.active.row_dimensions[1].height = 36

    column_width = {'A': 38, 'B': 14, 'C': 19.5, 'D': 58.5, 'E': 10, 'F': 10, 'G': 10}
    for column, width in column_width.items():
        report_inventory_template.active.column_dimensions[column].width = width

    report_inventory_template.active.merge_cells('A1:G1')
    report_inventory_template.active.merge_cells('A2:B2')
    for letter in 'CDEFG':
        report_inventory_template.active.merge_cells(f'{letter}2:{letter}3')

    report_inventory_template.active.cell(1, 1).font = dor.font_row_1
    report_inventory_template.active.cell(1, 1).alignment = dor.alignment
    for cell in report_inventory_template.active[2]:
        cell.font = dor.font_row_2
        cell.alignment = dor.alignment
    for cell in report_inventory_template.active[3]:
        cell.alignment = dor.alignment

    report_inventory_template.active.cell(
        1, 1, value=f'Инвентаризация {item} по складу {warehouse} за {datetime.today().strftime("%d.%m.%y")}')
    report_inventory_template.active.cell(2, 1, value="Место")
    report_inventory_template.active.cell(3, 1, value="Склад хранения")
    report_inventory_template.active.cell(3, 2, value="Склад отгрузки")
    report_inventory_template.active.cell(2, 3, value="Товар")
    report_inventory_template.active.cell(2, 4, value="Наименование")
    report_inventory_template.active.cell(2, 5, value="Остаток")
    report_inventory_template.active.cell(2, 6, value=int(warehouse) - 1)
    report_inventory_template.active.cell(2, 7, value=int(warehouse))

    return report_inventory_template


def create_report_refill_warehouse_template(warehouse):
    report_refill_warehouse_template = op.Workbook()
    report_refill_warehouse_template.remove(report_refill_warehouse_template.active)
    report_refill_warehouse_template.create_sheet(title=warehouse)

    report_refill_warehouse_template.active.row_dimensions[1].height = 36
    report_refill_warehouse_template.active.row_dimensions[2].height = 30

    column_width = {'A': 12, 'B': 50, 'C': 9, 'D': 9, 'E': 12, 'F': 12}
    for column, width in column_width.items():
        report_refill_warehouse_template.active.column_dimensions[column].width = width

    report_refill_warehouse_template.active.merge_cells('A1:F1')

    report_refill_warehouse_template.active.cell(1, 1).font = dor.font_row_1
    report_refill_warehouse_template.active.cell(1, 1).alignment = dor.alignment_center_wrap_text
    for cell in report_refill_warehouse_template.active[2]:
        cell.font = dor.font_row_2
        cell.alignment = dor.alignment_center_wrap_text

    report_refill_warehouse_template.active.cell(
        1, 1, value=f'Пополнение склада {warehouse} за {datetime.today().strftime("%d.%m.%y")}')
    report_refill_warehouse_template.active.cell(2, 1, value="Код товара")
    report_refill_warehouse_template.active.cell(2, 2, value="Наименование")
    report_refill_warehouse_template.active.cell(2, 3, value=f"Склад {int(warehouse) - 1} (шт.)")
    report_refill_warehouse_template.active.cell(2, 4, value=f"Склад {int(warehouse)} (шт.)")
    report_refill_warehouse_template.active.cell(2, 5, value="Ячейка хранения")
    report_refill_warehouse_template.active.cell(2, 6, value="Ячейка отгрузки")

    return report_refill_warehouse_template


def prepare_for_print(report):
    report.active.page_setup.orientation = report.active.ORIENTATION_LANDSCAPE
    report.active.page_setup.paperSize = report.active.PAPERSIZE_A4
    report.active.page_setup.fitToPage = True
    report.active.page_setup.fitToHeight = False
    report.active.page_margins = PageMargins(left=0.2, right=0.2, top=0.2, bottom=0.2)
    row = 1
    column_end = report.active.max_column
    values_of_row = [report.active.cell(row, col).value for col in range(1, column_end + 1)]
    while any(values_of_row):
        for i in range(1, column_end + 1):
            report.active.cell(row, i).border = dor.border
        row += 1
        values_of_row = [report.active.cell(row, col).value for col in range(1, column_end + 1)]
    return report
