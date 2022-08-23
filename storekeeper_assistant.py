import tkinter as tk
from datetime import datetime
from tkinter import filedialog as fd

import openpyxl as op
from openpyxl.styles import Side, Border

import mistake


class Application(tk.Tk):

    def __init__(self):
        super().__init__()
        screenwidth = self.winfo_screenwidth()
        screenheight = self.winfo_screenheight()
        app_win_width = 700
        app_win_height = 500
        x = (screenwidth - app_win_width) / 2
        y = (screenheight - app_win_height) / 2
        self.geometry("%dx%d+%d+%d" % (app_win_width, app_win_height, x, y))
        self.title('Помощник кладовщика')

        self.entry_file_path = tk.Entry()
        self.entry_file_path.insert(0, 'Вставьте файл отчета "Остатки АТ по партиям"')
        self.entry_file_path['width'] = 75
        self.entry_file_path.place(x=10, y=14)

        self.button_choose_file = tk.Button()
        self.button_choose_file.place(x=500, y=10)
        self.button_choose_file['text'] = 'Выберите файл'
        self.button_choose_file['font'] = 'Arial', 11
        self.button_choose_file['command'] = self.choose_file

        self.header = tk.Label()
        self.header['text'] = 'Провести следующие операции:'
        self.header['font'] = 'Arial', 12, 'italic'
        self.header.place(x=50, y=50)

        OPERATION_WAS_NOT_PERFORMED = {'Данная операция не выполнялась.': ''}
        self.MISTAKES_NOT_FOUND = {'Ошибки не обнаружены': ''}

        self.check_button_search_sku_in_two_places_value = tk.IntVar()
        self.check_button_search_sku_in_two_places = tk.Checkbutton(
            variable=self.check_button_search_sku_in_two_places_value,
            offvalue=0,
            onvalue=1)
        self.check_button_search_sku_in_two_places[
            'text'] = '  Найти товар, размещенный в двух и более ячейках отгрузки'
        self.check_button_search_sku_in_two_places.place(x=10, y=80)
        self.duplicates_result = OPERATION_WAS_NOT_PERFORMED

        self.check_button_search_two_or_more_sku_in_one_place_value = tk.IntVar()
        self.check_button_search_two_or_more_sku_in_one_place = tk.Checkbutton(
            variable=self.check_button_search_two_or_more_sku_in_one_place_value,
            offvalue=0,
            onvalue=1)
        self.check_button_search_two_or_more_sku_in_one_place[
            'text'] = '  Найти ячейки, в которых больше одного вида товара'
        self.check_button_search_two_or_more_sku_in_one_place.place(x=10, y=110)
        self.wrong_place = OPERATION_WAS_NOT_PERFORMED

        self.check_search_cells_empty_value = tk.IntVar()
        self.check_search_cells_empty = tk.Checkbutton(
            variable=self.check_search_cells_empty_value,
            offvalue=0,
            onvalue=1)
        self.check_search_cells_empty['text'] = '  Найти пустые нижние ячейки на складах отгрузки'
        self.check_search_cells_empty.place(x=10, y=140)
        self.cells_empty = OPERATION_WAS_NOT_PERFORMED

        self.header_2 = tk.Label()
        self.header_2['text'] = 'Если хотите сформировать отчет для пополнения склада отгрузки, \n поставьте галочку ' \
                                'и выберите файл отчета "Текущие остатки":'
        self.header_2['font'] = 'Arial', 12, 'italic'
        self.header_2.place(x=50, y=170)

        self.check_button_fill_storage_value = tk.IntVar()
        self.check_button_fill_storage = tk.Checkbutton(
            variable=self.check_button_fill_storage_value,
            offvalue=0,
            onvalue=1)
        self.check_button_fill_storage[
            'text'] = 'Сформировать отчет для пополнения склада отгрузки'
        self.check_button_fill_storage.place(x=10, y=220)

        self.unit_max = tk.StringVar()
        self.maximum = tk.Entry(textvariable=self.unit_max)
        self.maximum['width'] = 15
        self.maximum.place(x=10, y=250)

        self.header_max = tk.Label()
        self.header_max['text'] = ' Впишите максимальный остаток товара на складе отгрузки'
        self.header_max.place(x=50, y=250)

        self.entry_file_current_balance = tk.Entry()
        self.entry_file_current_balance.insert(0, 'Вставьте файл отчета "Текущие остатки (цена, шт., бр., пост.)"')
        self.entry_file_current_balance['width'] = 75
        self.entry_file_current_balance.place(x=10, y=300)

        self.button_choose_file_cb = tk.Button()
        self.button_choose_file_cb.place(x=500, y=296)
        self.button_choose_file_cb['text'] = 'Выберите файл'
        self.button_choose_file_cb['font'] = 'Arial', 11
        self.button_choose_file_cb['command'] = self.choose_file_cb

        self.button_run_actions = tk.Button()
        self.button_run_actions['text'] = 'Начать'
        self.button_run_actions['font'] = 'Arial', 11
        self.button_run_actions['command'] = self.run_actions
        self.button_run_actions.place(x=50, y=350)

    # Выбор файла 'Остатки АТ по партиям'
    def choose_file(self):
        self.file_name_bb = fd.askopenfilename(filetypes=(('Excel files', '.xlsx'),))
        self.entry_file_path.delete(0, 'end')
        self.entry_file_path.insert(0, self.file_name_bb)

    # Выбор файла 'Текущие остатки'
    def choose_file_cb(self):
        self.file_name_cb = fd.askopenfilename(filetypes=(("Excel files", ".xlsx"),))
        self.entry_file_current_balance.delete(0, 'end')
        self.entry_file_current_balance.insert(0, self.file_name_cb)

    def open_file(self, file_name):
        sheet = op.load_workbook(file_name).sheetnames[0]
        if sheet == 'tmp':
            self.file_balance_current = op.load_workbook(file_name)  # Открытие файла 'Текущие остатки'
            self.file_path_cb = file_name
        else:
            self.file_balance_batch = op.load_workbook(file_name)  # Открытие файла 'Остатки АТ по партиям

    def run_actions(self):
        self.open_file(self.file_name_bb)  # Остатки АТ по партиям
        self.open_file(self.file_name_cb)  # Текущие остатки
        if self.check_button_search_sku_in_two_places_value.get():
            self.duplicates_result = mistake.search_sku_in_two_places(self.file_balance_batch)
        if self.check_button_search_two_or_more_sku_in_one_place_value.get():
            self.wrong_place = mistake.search_two_or_more_sku_in_one_place(self.file_balance_batch)
        if self.check_search_cells_empty_value.get():
            self.cells_empty = mistake.search_cells_empty(self.file_balance_batch)
        if self.check_button_fill_storage_value.get():
            # файл 'текущие остатки'
            report_406 = self.file_balance_current
            self.create_report(report_406, '406')
            report_437 = op.load_workbook(self.file_path_cb)
            self.create_report(report_437, '437')
        self.fill_errors()
        self.destroy()

    def create_report(self, report, warehouse):
        if warehouse == '406':
            report.active.delete_cols(19, 16)
            report.active.delete_cols(4, 13)
        else:
            report.active.delete_cols(21, 14)
            report.active.delete_cols(4, 15)

        report = self.prepare_file(report)
        report = self.fill_report(report)
        report.save(f'Отчет от {datetime.today().strftime("%d.%m.%y %H_%M_%S")} ({warehouse}).xlsx')

    # prepare file for preview
    def prepare_file(self, file):
        file.active.delete_cols(1)
        file.active.delete_rows(5)
        # удаление позиций, отсутствующих на складе хранения
        # отбор позиций для пополнения склада отгрузки
        row = 6
        col = 1
        sku = file.active.cell(row=row, column=col).value
        balance_min = int(self.maximum.get())
        while isinstance(sku, str):
            amount_store = file.active.cell(row=row, column=3).value  # кол-во на складе хранения
            amount_sell = file.active.cell(row=row, column=4).value  # кол-во на складе отгрузки
            if amount_sell is None:
                amount_sell = 0
            if amount_store is None or amount_sell > balance_min:
                file.active.delete_rows(row)
                row -= 1
            row += 1
            sku = file.active.cell(row=row, column=col).value
        # ширина столбцов, отрисовка границ ячеек
        for i in 'AEF':
            file.active.column_dimensions[i].width = 12
        file.active.column_dimensions['B'].width = 50
        file.active.column_dimensions['C'].width = 9
        file.active.column_dimensions['D'].width = 9
        rows = file.active.max_row - 5
        row = 5
        side = Side(border_style="thin", color="000000")
        for _ in range(rows):
            file.active.row_dimensions[row].height = 15
            file.active.cell(row=row, column=5).border = Border(left=side, right=side, top=side,
                                                                bottom=side)
            file.active.cell(row=row, column=6).border = Border(left=side, right=side, top=side,
                                                                bottom=side)
            row += 1
        return file

    def fill_report(self, file):
        # файл 'остатки АТ по партиям'
        cells_file = self.file_balance_batch

        # список sku для пополнения склада отгрузки
        sku_list = []
        row_405 = 6
        sku_405 = file.active.cell(row_405, 1).value
        while isinstance(sku_405, str):
            sku_list.append(sku_405)
            row_405 += 1
            sku_405 = file.active.cell(row_405, 1).value

        sku_amount = len(sku_list)
        row_406 = 4
        # sku_report - код товара, который надо пополнить
        # sku_batch_curr - код товара (рассматриваемый) в файле "Остатки по партиям"
        # sku_batch_next - код товара (следующий) в файле "Остатки по партиям"
        # sku_store_405 - ячейка на складе хранения (откуда взять)
        # sku_store_406 - ячейка на складе отгрузки (куда поставить)
        # cells_405 - ячейки на складе хранения
        # cells_406 - ячейки на складе отгрузки
        # row_406 - номер строки в файле "Остатки по партиям"
        cells_405_set = []
        cells_406_set = []
        sku_batch_curr = cells_file.active.cell(row_406, 4).value
        while sku_batch_curr is not None:
            sku_batch_curr = sku_batch_curr.split('.')[0]
            if sku_batch_curr not in sku_list:
                row_406 += 1
                sku_batch_curr = cells_file.active.cell(row_406, 4).value
            else:
                cells_405_list = cells_file.active.cell(row_406, 1).value
                if cells_405_list is not None:
                    cells_405_set = self.create_cells_list(cells_405_list, cells_405_set)
                cells_406_list = cells_file.active.cell(row_406, 2).value
                if cells_406_list is not None:
                    cells_406_set = self.create_cells_list(cells_406_list, cells_406_set)
                sku_batch_next = cells_file.active.cell(row_406 + 1, 4).value
                if sku_batch_next is not None:
                    sku_batch_next = sku_batch_next.split('.')[0]
                    if sku_batch_next != sku_batch_curr:
                        row_report = 6
                        for _ in range(sku_amount):
                            if sku_batch_curr == file.active.cell(row_report, 1).value:
                                if not cells_405_set:
                                    cells_405_set = ['пусто']
                                file.active.cell(row_report, 5).value = cells_405_set[0]
                                if not cells_406_set:
                                    cells_406_set = ['пусто']
                                file.active.cell(row_report, 6).value = ", ".join(cells_406_set)
                                cells_405_set = []
                                cells_406_set = []
                                break
                            else:
                                row_report += 1
                else:
                    row_report = 6
                    for _ in range(sku_amount):
                        if sku_batch_curr == file.active.cell(row_report, 1).value:
                            file.active.cell(row_report, 5).value = cells_405_set[0]
                            file.active.cell(row_report, 6).value = ", ".join(cells_406_set)
                            cells_405_set = []
                            cells_406_set = []
                            break
                        else:
                            row_report += 1
                sku_batch_curr = sku_batch_next
                row_406 += 1
        return file

    def create_cells_list(self, cells_list, cells_set):
        cells_list = cells_list.split(', ')
        for cell in cells_list:
            if cell not in cells_set:
                cells_set.append(cell)
        return cells_set

    def fill_errors(self):
        row = 1
        col = 1
        result = op.Workbook()
        result.remove(result.active)
        sheet_1 = result.create_sheet('1')
        sheet_1.cell(row=row, column=col).value = 'Товар, размещенный в двух и более ячейках отгрузки'
        if not self.duplicates_result:
            self.duplicates_result = self.MISTAKES_NOT_FOUND
        for k, v in self.duplicates_result.items():
            row += 1
            sheet_1.cell(row=row, column=col).value = k
            sheet_1.cell(row=row, column=col + 1).value = ', '.join(v)

        row += 2
        sheet_1.cell(row=row, column=col).value = 'Ячейки, в которых больше одного вида товара'
        if not self.wrong_place:
            self.wrong_place = self.MISTAKES_NOT_FOUND
        for key, value in self.wrong_place.items():
            row += 1
            sheet_1.cell(row=row, column=col).value = key
            sheet_1.cell(row=row, column=col + 1).value = value

        row += 2
        sheet_1.cell(row=row, column=col).value = 'Свободные ячейки'
        row += 1
        sheet_1.cell(row=row, column=col).value = 406
        sheet_1.cell(row=row, column=col + 1).value = 437
        row_406 = row_437 = row + 1
        if self.check_search_cells_empty_value.get() == 0:
            sheet_1.cell(row=row_406, column=col).value = next(iter(self.cells_empty))
        else:
            for c_406 in self.cells_empty[0]:
                sheet_1.cell(row=row_406, column=col).value = c_406
                row_406 += 1
            for c_437 in self.cells_empty[1]:
                sheet_1.cell(row=row_437, column=col + 1).value = c_437
                row_437 += 1

        result.save(f'Ошибки от {datetime.today().strftime("%d.%m.%y %H_%M_%S")}.xlsx')


app = Application()
app.mainloop()
