# all functions for foreman
import openpyxl as op
from openpyxl.worksheet.page import PageMargins
from datetime import datetime

import warehouses as w
import design_of_report as dor


def create_report_template(warehouse, item):
    report_template = op.Workbook()
    report_template.remove(report_template.active)
    report_template.create_sheet(title=warehouse)

    report_template.active.row_dimensions[1].height = 36

    column_width = {'A': 38, 'B': 14, 'C': 19.5, 'D': 58.5, 'E': 10, 'F': 10, 'G': 10}
    for column, width in column_width.items():
        report_template.active.column_dimensions[column].width = width

    report_template.active.merge_cells('A1:G1')
    report_template.active.merge_cells('A2:B2')
    for letter in 'CDEFG':
        report_template.active.merge_cells(f'{letter}2:{letter}3')

    for cell in report_template.active[2]:
        cell.font = dor.font_row_2
        cell.alignment = dor.alignment

    for cell in report_template.active[3]:
        cell.alignment = dor.alignment

    report_template.active.cell(
        1, 1, value=f'Инвентаризация {item} по складу {warehouse} за {datetime.today().strftime("%d.%m.%y")}')
    report_template.active.cell(1, 1).font = dor.font_row_1
    report_template.active.cell(1, 1).alignment = dor.alignment

    report_template.active.cell(2, 1, value="Место")
    report_template.active.cell(3, 1, value="Склад хранения")
    report_template.active.cell(3, 2, value="Склад отгрузки")
    report_template.active.cell(2, 3, value="Товар")
    report_template.active.cell(2, 4, value="Наименование")
    report_template.active.cell(2, 5, value="Остаток")
    report_template.active.cell(2, 6, value=int(warehouse) - 1)
    report_template.active.cell(2, 7, value=int(warehouse))

    return report_template


def prepare_for_print(report_inventory):
    report_inventory.active.page_setup.orientation = report_inventory.active.ORIENTATION_LANDSCAPE
    report_inventory.active.page_setup.paperSize = report_inventory.active.PAPERSIZE_A4
    report_inventory.active.page_setup.fitToPage = True
    report_inventory.active.page_setup.fitToHeight = False
    report_inventory.active.page_margins = PageMargins(left=0.2, right=0.2, top=0.2, bottom=0.2)
    row = 1
    values_of_row = [report_inventory.active.cell(row, col).value for col in range(1, 8)]
    while any(values_of_row):
        for i in range(1, 8):
            report_inventory.active.cell(row, i).border = dor.border
        row += 1
        values_of_row = [report_inventory.active.cell(row, col).value for col in range(1, 8)]
    return report_inventory


def add_row_in_inventory(file_AT, row, column_warehouse):
    values_of_row = [file_AT.active.cell(row, 1).value, file_AT.active.cell(row, 2).value,
                     file_AT.active.cell(row, 4).value, file_AT.active.cell(row, 5).value,
                     file_AT.active.cell(row, 6).value, file_AT.active.cell(row, column_warehouse - 1).value,
                     file_AT.active.cell(row, column_warehouse).value]
    return values_of_row


def create_file_inventory_cells(file_AT, warehouse, cell_start, cell_finish):
    report_inventory_cells = create_report_template(warehouse, 'ячеек')
    cells_in_warehouse = []
    values_of_second_row = [str(file_AT.active.cell(2, column).value) for column in range(1, 11)]
    column_warehouse = values_of_second_row.index(warehouse) + 1  # номер столбца с номером склада, который считают
    row_start = 4  # с 4-ой строки начинаются данные
    rows = file_AT.active.max_row
    if warehouse == w.warehouse_406:
        cells_in_warehouse = w.cells_all_406
    elif warehouse == w.warehouse_437:
        cells_in_warehouse = w.cells_all_437
    index_start = cells_in_warehouse.index(cell_start)
    index_finish = cells_in_warehouse.index(cell_finish) + 1
    cells_for_inventory = cells_in_warehouse[index_start:index_finish]
    values_of_cells_for_inventory = []  # список списков со значениями из строк с нужными ячейками
    for _ in range(rows - row_start + 1):
        target_cells = file_AT.active.cell(row_start, 2).value
        if target_cells is not None and len(target_cells) < 10 and target_cells in cells_for_inventory:
            values_of_row = add_row_in_inventory(file_AT, row_start, column_warehouse)
            values_of_cells_for_inventory.append(values_of_row)
            row_start += 1
        elif target_cells is not None and len(target_cells) > 9:
            target_cells.split(', ')
            for t in target_cells:
                if t in cells_for_inventory:
                    values_of_row = add_row_in_inventory(file_AT, row_start, column_warehouse)
                    values_of_cells_for_inventory.append(values_of_row)
            row_start += 1
        else:
            row_start += 1
    row_start = 4
    for value in values_of_cells_for_inventory:
        if value[0] is not None:
            height = (int(len(value[0]) / 41) + 1) * 15
            report_inventory_cells.active.row_dimensions[row_start].height = height
            report_inventory_cells.active.cell(row_start, 1).alignment = dor.alignment_wrap_text
        for c in range(1, 8):
            report_inventory_cells.active.cell(row_start, c).value = value[c - 1]
        row_start += 1
    report_inventory_cells = prepare_for_print(report_inventory_cells)
    report_inventory_cells.save(f'Инвентаризация ячеек {datetime.today().strftime("%d.%m.%y %H_%M_%S")} ({warehouse}).xlsx')

# file_supervisor - файл с ошибками
def create_file_inventory_supervisor(file_AT, file_supervisor, warehouse, date_start, date_finish):
    report_inventory_mistakes = create_report_template(warehouse, 'ошибок')
    values_of_second_row = [str(file_AT.active.cell(2, column).value) for column in range(1, 11)]
    column_warehouse = values_of_second_row.index(warehouse) + 1  # номер столбца с номером склада, который считают
    sheet_active = file_supervisor["СБОРКА"]
    file_supervisor.active = sheet_active
    rows = sheet_active.max_row
    sku_mistake = []
    for i in range(rows, 0, -1):
        date_mistake = sheet_active.cell(i, 2).value.date()
        sku = sheet_active.cell(i, 3).value
        if date_start <= date_mistake <= date_finish:
            if sku is not None:
                sku_mistake.append(str(sku))
        else:
            break
    values_of_cells_for_inventory = []  # список списков со значениями из строк с нужными ячейками
    row_start = 4
    sku_at = file_AT.active.cell(row_start, 4).value
    while sku_at is not None:
        sku_at = sku_at.split('.')[0]
        if sku_at in sku_mistake and file_AT.active.cell(row_start, column_warehouse).value > 0:
            values_of_row = add_row_in_inventory(file_AT, row_start, column_warehouse)
            values_of_cells_for_inventory.append(values_of_row)
        row_start += 1
        sku_at = file_AT.active.cell(row_start, 4).value
    row_start = 4
    for value in values_of_cells_for_inventory:
        if value[0] is not None:
            height = (int(len(value[0]) / 41) + 1) * 15
            report_inventory_mistakes.active.row_dimensions[row_start].height = height
            report_inventory_mistakes.active.cell(row_start, 1).alignment = dor.alignment_wrap_text
        for c in range(1, 8):
            report_inventory_mistakes.active.cell(row_start, c).value = value[c - 1]
        row_start += 1
    report_inventory_mistakes = prepare_for_print(report_inventory_mistakes)
    report_inventory_mistakes.save(
        f'Инвентаризация ошибок {datetime.today().strftime("%d.%m.%y %H_%M_%S")} ({warehouse}).xlsx')
